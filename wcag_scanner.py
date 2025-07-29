#!/usr/bin/env python3
"""
WCAG Automated Scanner
======================

This script provides a simple command‑line tool to perform a basic accessibility scan
on a web page or local HTML file.  It was inspired by the features offered in
popular open‑source accessibility scanners such as Pa11y, AInspector for Firefox
and Deque's axe‑core engine.  Those tools allow testers to choose a WCAG
conformance level, run automated rules against a page and export the results
in formats like CSV or JSON【249372618707913†L65-L97】.  They also support running
on the command line and generating reports in spreadsheets【249372618707913†L100-L127】.

Because this environment cannot install external Node packages at run time, this
implementation uses Python's standard libraries together with BeautifulSoup to
parse the markup and perform a handful of WCAG 2.2 Level A/AA checks.  It is not
a replacement for full rule engines like axe‑core—which includes rules for
WCAG 2.0, 2.1 and 2.2 across levels A, AA and AAA【299890540444768†L178-L182】—but it
demonstrates how an automated scan might work and how the results can be
written into a spreadsheet.  If you wish to test with a full rule set, you can
swap out the scanning logic for calls to an open‑source engine such as pa11y
(`pa11y <url> --runner axe --reporter json`) and parse its JSON output into
a DataFrame.

Usage
-----

Run the script with a URL or a path to a local HTML file:

```
python wcag_scanner.py --url https://www.example.com
python wcag_scanner.py --file path/to/page.html
```

The script fetches the markup, runs a small set of accessibility checks and
writes an Excel file (`report.xlsx`) with the results.  A custom
"Perspective Tester" logo is inserted at the top of the spreadsheet to brand
the report.
"""

import argparse
import os
from typing import List, Dict, Optional
import requests
from bs4 import BeautifulSoup
import pandas as pd
from pandas import ExcelWriter
from openpyxl.drawing.image import Image as XLImage
import openpyxl  # used for column letter conversion


def fetch_html_from_url(url: str) -> Optional[str]:
    """Fetch HTML content from a remote URL.

    Args:
        url: The URL to fetch.

    Returns:
        A string containing the HTML, or ``None`` on failure.
    """
    try:
        response = requests.get(url, timeout=30)
        # If the server blocks requests (returns 403), instruct the user.
        if response.status_code != 200:
            print(
                f"Warning: received status code {response.status_code} from {url}.\n"
                "The script will continue but you might need to provide a local HTML file."
            )
        return response.text
    except Exception as exc:
        print(f"Error fetching {url}: {exc}")
        return None


def read_local_file(path: str) -> Optional[str]:
    """Read HTML content from a local file."""
    if not os.path.isfile(path):
        print(f"File not found: {path}")
        return None
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as fh:
            return fh.read()
    except Exception as exc:
        print(f"Error reading {path}: {exc}")
        return None


def check_images_have_alt(soup: BeautifulSoup) -> List[Dict[str, str]]:
    """Check that all <img> elements include non‑empty alt attributes.

    Returns a list of issue dictionaries.
    """
    issues = []
    for img in soup.find_all('img'):
        alt = img.get('alt')
        if alt is None or alt.strip() == '':
            # Build a CSS selector approximation using the tag's position.
            # We limit the selector to nth-of-type within the parent for readability.
            try:
                index = list(img.parent.find_all('img')).index(img) + 1
            except ValueError:
                index = 1
            selector = f"img:nth-of-type({index})"
            issues.append({
                'rule': 'IMG_ALT',
                'wcag_principle': 'Perceivable',
                'wcag_success_criterion': '1.1.1 Non-text Content',
                'level': 'A',
                'description': 'Image elements must have a text alternative (alt attribute).',
                'context': str(img)[:200],
                'selector': selector,
                'recommendation': 'Add a descriptive alt attribute to the <img> element.'
            })
    return issues


def check_form_inputs_have_labels(soup: BeautifulSoup) -> List[Dict[str, str]]:
    """Check that form controls have associated labels.

    For inputs, selects and textareas, a control should either have a <label> element
    referencing its id or wrap the control.  Controls of type hidden, submit and button
    are ignored.  The check maps to WCAG 2.2 success criterion 3.3.2 Labels or
    Instructions (Level A) and 1.3.1 Info and Relationships (Level A).
    """
    issues = []
    form_controls = soup.find_all(['input', 'select', 'textarea'])
    for control in form_controls:
        if control.name == 'input' and control.get('type') in ['hidden', 'submit', 'button', 'image']:
            continue
        # Determine if the control has an associated label
        has_label = False
        # Check if wrapped by a <label> ancestor
        for parent in control.parents:
            if parent.name == 'label':
                has_label = True
                break
        # Check for <label for="id"> association
        control_id = control.get('id')
        if not has_label and control_id:
            label = soup.find('label', attrs={'for': control_id})
            if label:
                has_label = True
        if not has_label:
            try:
                index = list(control.parent.find_all(control.name)).index(control) + 1
            except ValueError:
                index = 1
            selector = f"{control.name}:nth-of-type({index})"
            issues.append({
                'rule': 'FORM_LABEL',
                'wcag_principle': 'Understandable',
                'wcag_success_criterion': '3.3.2 Labels or Instructions / 1.3.1 Info and Relationships',
                'level': 'A',
                'description': 'Form controls must have associated labels.',
                'context': str(control)[:200],
                'selector': selector,
                'recommendation': 'Add a <label> element referencing the control or wrap the control in a <label>.'
            })
    return issues


def check_links_have_descriptive_text(soup: BeautifulSoup) -> List[Dict[str, str]]:
    """Check that link text is meaningful when read out of context.

    This addresses WCAG 2.2 success criterion 2.4.4 Link Purpose (In Context) at
    Level A and 2.4.9 Link Purpose (Link Only) at Level AAA.  The implementation
    warns when link text is generic (e.g. 'click here', 'more', 'read more').
    """
    generic_phrases = ['click here', 'more', 'read more', 'learn more', 'details']
    issues = []
    for link in soup.find_all('a', href=True):
        text = link.get_text().strip().lower()
        if not text or any(text == phrase or text.startswith(phrase + ' ') for phrase in generic_phrases):
            try:
                index = list(link.parent.find_all('a')).index(link) + 1
            except ValueError:
                index = 1
            selector = f"a:nth-of-type({index})"
            issues.append({
                'rule': 'LINK_TEXT',
                'wcag_principle': 'Understandable',
                'wcag_success_criterion': '2.4.4 Link Purpose (In Context)',
                'level': 'A',
                'description': 'Link text should be descriptive and convey the purpose of the link.',
                'context': str(link)[:200],
                'selector': selector,
                'recommendation': 'Replace generic link text with text that describes the destination or action.'
            })
    return issues


def run_accessibility_checks(html: str) -> List[Dict[str, str]]:
    """Run a series of accessibility checks against HTML.

    Returns a list of issue dictionaries.
    """
    soup = BeautifulSoup(html, 'html.parser')
    issues: List[Dict[str, str]] = []
    issues += check_images_have_alt(soup)
    issues += check_form_inputs_have_labels(soup)
    issues += check_links_have_descriptive_text(soup)
    return issues


def write_report(issues: List[Dict[str, str]], output_path: str, logo_path: str,
                 url: Optional[str] = None) -> None:
    """Write the collected issues to an Excel spreadsheet with branding and a summary sheet.

    The generated workbook contains two sheets:

    * ``Status`` – a high‑level overview for the scanned URL showing the count of
      issues by rule and a pass/fail status for each category.  The header row
      is styled with a coloured fill similar to the provided example file.  If
      a category has zero issues, it is marked ``Pass`` and coloured green.
      Otherwise it is marked ``Fail`` and coloured red.  The total number of
      issues is also reported.
    * ``Issues`` – a detailed listing of every issue found.  The header row is
      styled with a different background colour to separate it visually from
      the data rows.  Columns include the rule name, WCAG principle,
      success criterion, level, description, context, selector and
      recommendation.

    Args:
        issues: List of issue dictionaries collected from the scan.
        output_path: Path for the Excel file to be created.
        logo_path: Path to the Perspective Tester logo image used for branding.
        url: The URL that was scanned.  If provided, it will be displayed on
            the summary sheet.  When scanning a local file, this may be
            omitted.
    """
    # Organise the issues into a DataFrame for convenience
    df = pd.DataFrame(issues)
    expected_cols = [
        'rule',
        'wcag_principle',
        'wcag_success_criterion',
        'level',
        'description',
        'context',
        'selector',
        'recommendation',
    ]
    for col in expected_cols:
        if col not in df.columns:
            df[col] = ''
    df = df[expected_cols]

    # Determine counts per rule for the summary
    rule_counts = {
        'Missing Alt Text': int((df['rule'] == 'IMG_ALT').sum()),
        'Unlabelled Form Controls': int((df['rule'] == 'FORM_LABEL').sum()),
        'Generic Link Text': int((df['rule'] == 'LINK_TEXT').sum()),
    }

    total_issues = len(issues)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = 'Status'

    # Insert logo on summary sheet
    if logo_path and os.path.isfile(logo_path):
        try:
            img = XLImage(logo_path)
            # Resize to fit nicely at the top left
            img.width = 200
            img.height = 80
            ws_summary.add_image(img, 'A1')
        except Exception as exc:
            print(f"Could not insert logo: {exc}")

    # Header for summary table
    header = ['URL', 'Missing Alt Text', 'Unlabelled Form Controls', 'Generic Link Text', 'Total Issues']
    ws_summary.append([])  # leave first row (logo)
    ws_summary.append([])
    ws_summary.append([])
    start_row = ws_summary.max_row + 1

    # Apply styling to header row (light fill and bold)
    header_fill = PatternFill(fill_type='solid', fgColor='FFE6B8AF')  # pale peach similar to sample
    # Define a thin border to apply around each cell, similar to the example file
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, col_name in enumerate(header, start=1):
        cell = ws_summary.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border

    # Data row for summary
    url_display = url or 'N/A'
    data = [url_display, rule_counts['Missing Alt Text'], rule_counts['Unlabelled Form Controls'],
            rule_counts['Generic Link Text'], total_issues]
    # Create data row immediately following the header
    data_row_index = start_row + 1
    # Fill URL cell
    url_cell = ws_summary.cell(row=data_row_index, column=1)
    url_cell.value = url_display
    url_cell.alignment = Alignment(horizontal='left', vertical='center')
    url_cell.border = thin_border
    # Style data row: determine pass/fail for each rule count and assign colours
    pass_fill = PatternFill(fill_type='solid', fgColor='FF92D050')  # light green
    fail_fill = PatternFill(fill_type='solid', fgColor='FFFFC000')  # light amber
    rule_values = [rule_counts['Missing Alt Text'], rule_counts['Unlabelled Form Controls'], rule_counts['Generic Link Text']]
    for idx, value in enumerate(rule_values, start=2):
        cell = ws_summary.cell(row=data_row_index, column=idx)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        if value == 0:
            cell.value = 'Pass'
            cell.fill = pass_fill
        else:
            cell.value = 'Fail'
            cell.fill = fail_fill
    # Total issues cell
    tot_cell = ws_summary.cell(row=data_row_index, column=len(header))
    tot_cell.value = total_issues
    tot_cell.alignment = Alignment(horizontal='center', vertical='center')
    tot_cell.border = thin_border

    # Adjust column widths based on header length
    for col_idx, col_name in enumerate(header, start=1):
        ws_summary.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(15, len(col_name) + 5)

    # Detailed issues sheet
    ws_details = wb.create_sheet('Issues')
    # Insert logo again on details sheet to mirror sample aesthetic
    if logo_path and os.path.isfile(logo_path):
        try:
            img2 = XLImage(logo_path)
            img2.width = 200
            img2.height = 80
            ws_details.add_image(img2, 'A1')
        except Exception:
            pass
    # Leave a couple of blank rows under the logo
    ws_details.append([])
    ws_details.append([])
    # Write header row for issues table
    issue_header_fill = PatternFill(fill_type='solid', fgColor='FFD9E2F3')  # light purple similar to example
    # Create header row for details sheet
    ws_details.append(expected_cols)
    header_row = ws_details.max_row
    for col_idx, col_name in enumerate(expected_cols, start=1):
        cell = ws_details.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = issue_header_fill
        cell.border = thin_border
    # Append issue rows
    for _, row in df.iterrows():
        ws_details.append(list(row))
    # Style issue rows (borders and alignment)
    for row_idx in range(header_row + 1, ws_details.max_row + 1):
        for col_idx in range(1, len(expected_cols) + 1):
            cell = ws_details.cell(row=row_idx, column=col_idx)
            # Left align for most columns, except maybe the first numeric ones center align
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border
    # Adjust column widths for details sheet
    for col_idx, col_name in enumerate(expected_cols, start=1):
        series = df[col_name].astype(str).head(50)
        max_len = max(len(str(col_name)), max(series.apply(len)) if not series.empty else 0)
        ws_details.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 5, 50)

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description='Run a basic WCAG accessibility scan.')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--url', help='URL of the page to scan')
    group.add_argument('--file', help='Local HTML file to scan')
    parser.add_argument('--output', default='report.xlsx', help='Output Excel file name')
    parser.add_argument('--logo', default='logo.png', help='Path to the Perspective Tester logo image')
    args = parser.parse_args()

    if args.url:
        html = fetch_html_from_url(args.url)
    else:
        html = read_local_file(args.file)
    if not html:
        print('No HTML content was retrieved. Exiting.')
        return
    issues = run_accessibility_checks(html)
    # Use the URL or filename for the summary sheet if provided
    summary_identifier: Optional[str] = args.url if args.url else args.file
    write_report(issues, args.output, args.logo, url=summary_identifier)
    print(f'Report generated: {args.output} ({len(issues)} issues found)')


if __name__ == '__main__':
    main()